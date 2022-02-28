# function for analysis of runoff timeseries
# Mohammad Areeb Anwer

import os
import pandas as pd
from matplotlib import pyplot as plt
import openpyxl
from openpyxl import load_workbook


def annual_average(df_monthly, file_name):
    """
        Function for calulating Annual average from monthly data, plotting graph and saving in Excel workbook.
        :param df_monthly: Dataframe - contains monthly data
        :param file_name: STR - file name containing respective coordinates
        :output: save excel workbook with annual average, graph and respective name
    """

    annual_avg = df_monthly.groupby(pd.Grouper(freq='A', dropna= False)).mean()  # Annual Average
    filename_xlsx = file_name + ".xlsx"  # file name with extension
    book = load_workbook(os.path.abspath("..") + "\\GRUN_reader\\output\\" + filename_xlsx)  # load workbook
    writer = pd.ExcelWriter(os.path.abspath("..") + "\\GRUN_reader\\output\\" + filename_xlsx, engine='openpyxl')
    writer.book = book
    annual_avg.to_excel(writer, sheet_name='Annual_avg')  # Save DataFrame to worksheet with name
    writer.save()
    # Create and save Graph
    plt.style.use('seaborn')  # Plot style
    years = annual_avg.index  # years array x-axis
    runoff = annual_avg['Runoff']  # Runoff array y-axis
    plt.figure(figsize=(12, 3))  # dimensions of Graph
    plt.scatter(years, runoff)  # Graph type scatter
    plt.title(file_name[:-7] + ' Annual mean Runoff')  # graph title
    plt.xlabel('Years')  # x-axis label
    plt.ylabel('Runoff(cumecs)')  # y-axis label
    plt.tight_layout()  # to fit graph in extents
    graph_file = file_name + ".png"  # file name with extension
    plt.savefig(os.path.abspath("..") + "\\GRUN_reader\\output\\" + graph_file)  # save graph
    worksheet = writer.sheets['Annual_avg']  # calling respective worksheet
    graph = openpyxl.drawing.image.Image(os.path.abspath("..") + "\\GRUN_reader\\output\\" + graph_file)  # save graph as openpyxl object
    graph.anchor = 'C1'  # cell to add graph in
    worksheet.add_image(graph)  # add graph in respective worksheet
    writer.save()
    os.remove(os.path.abspath("..") + "\\GRUN_reader\\output\\" + graph_file)  # delete graph image from folder


def annual_maximum(df_monthly, file_name):
    """
        Function for calulating Annual average from monthly data, plotting graph and saving in Excel workbook.
        :param df_monthly: Dataframe - contains monthly data
        :param file_name: STR - file name containing respective coordinates
        :output: save excel workbook with annual average, graph and respective name
    """
    annual_max = df_monthly.loc[df_monthly.groupby(pd.Grouper(freq='A')).idxmax().iloc[:, 0]]  # calculate annual max
    filename_xlsx = file_name + ".xlsx"
    book = load_workbook(os.path.abspath("..") + "\\GRUN_reader\\output\\" + filename_xlsx)
    writer = pd.ExcelWriter(os.path.abspath("..") + "\\GRUN_reader\\output\\" + filename_xlsx, engine='openpyxl')
    writer.book = book
    annual_max.to_excel(writer, sheet_name='Annual_max')
    writer.save()
    # Create and save Graph
    plt.style.use('seaborn')
    years = annual_max.index
    runoff = annual_max['Runoff']
    plt.figure(figsize=(12, 3))
    plt.scatter(years, runoff)
    plt.title(file_name[:-7] + ' Annual Maximum Runoff')
    plt.xlabel('Years')
    plt.ylabel('Runoff(cumecs)')
    plt.tight_layout()
    graph_file = file_name + ".png"
    plt.savefig(os.path.abspath("..") + "\\GRUN_reader\\output\\" + graph_file)
    worksheet = writer.sheets['Annual_max']
    graph = openpyxl.drawing.image.Image(os.path.abspath("..") + "\\GRUN_reader\\output\\" + graph_file)
    graph.anchor = 'C1'
    worksheet.add_image(graph)
    writer.save()
    os.remove(os.path.abspath("..") + "\\GRUN_reader\\output\\" + graph_file)














